# As was mentioned in the project description of my portfolio, I created this project to speed up a reporting process by
# about 40min. The process begins with receiving data via emails, updates essential workbooks of that data, and then
# uploading those essential workbooks to the cloud. I will attempt to comment as best I can about what each step is
# doing, however, the code itself is deliberately vague for privacy.

import win32com.client
from openpyxl import load_workbook
from win32com.client import Dispatch
from subprocess import Popen
from datetime import datetime, timedelta
import os
import time

# First, we define the correct day that the data reflects
correct_date = input("The data date gets input here: ")

# Next, we define the prior business day to catpure day-over-day changes in data for various workbooks
old_date = input("The prior day is placed here:")

# Lastly, we define the day prior to the previous day. This is because, in order to capture correct day-over-day changes
# we must replace the prior previous day with the new previous day.
older_date = input("The day before the prior day is placed here:")

# Our program begins with instantiating an Outlook object
outlook = win32com.client.Dispatch('Outlook.Application').GetNameSpace("MAPI")

# Next, we define the inbox of our email account
inbox = outlook.GetDefaultFolder(6)

# We then perform a filtering out process to extract the correct emails that contain the relevant information
messages = inbox.Items  # Instantiate a message object for the messages in our inbox
date = datetime.now() - timedelta(days=1)  # Exclude all emails that were sent any time before the previous 24hrs
date = date.strftime('%m/%d/%y')  # Format the date as indicated
messages = messages.Restrict("[ReceivedTime] >= '" + date + "'")  # Apply our time filter to the inbox
messages = messages.Restrict("[Subject] = 'Subject Line of Email'")  # Extract the correct email based on the subject

# There are numerous emails that get sent out containing updated data. As a result, I repeat the process below for each
# email in order to extract the data attachment
for message in messages:  # Iterate through the remaining emails after our filter
    attachments = message.Attachments  # Instantiate an attachments object
    attachment = attachments.Item(1)  # Define the attachment
    attachment.SaveAsFile(os.path.join(r'file path', f'name of file {correct_date}.filetype'))  # Save the attachment

# After all of our data has been extracted, the first item of business is to update the date in a specific file. This
# date is referenced by several other files, so it is crucial that this file reflect the correct date.
file1_path = r'file\path'  # Define the file path
file1_wb = load_workbook(file1_path)  # Create an instance of the file
file1_ws = file1_wb.active  # Define the correct sheet of the file
format_date = datetime.strptime(correct_date, '%m/%d/%y')  # Parse the date with the format provided
file1_ws['CellNumber'].value = format_date  # Place the date from the beginning of the program into the file
file1_wb.save(file1_path)  # Save the file
file1_wb.close()  # Close the file

# Similar to the emails, I repeat the process below for several files. From a data file that we saved earlier, we copy
# and paste those values into another file that is referenced by other workbooks. This ensures that the other workbooks
# referencing this so-called "main" file are have the most recent data.
file2_path = r'file2path'  # Define the main file path
temp_file2_path = r'temp_file2path'  # Define the path to data sent over in our email that we saved previously

# We then create an instance of our files
file2_wb = load_workbook(file2_path)
temp_wb = load_workbook(temp_file2_path)

# We then define the correct sheets
file2_ws = file2_wb.active
temp_ws = temp_wb.active

# Due to formatting issues, a simple copy/paste function does not work. As a result, we perform a list comprehension to
# append all the value from the temporary files to a list. We then iterate through that list and place each data entry
# into the corresponding cell in the "main" file. While this only showcases a single column of data being updated, I
# perform the process below a plethora of times to update the entire workbook, changing the column values each time.
colLetter = [temp_ws.cell(row=i, column=123).value for i in range(1, 123456)]
for i, value in enumerate(colLetter):
    file2_ws.cell(row=123 + i, column=123, value=value)

# Save and close these two workbooks
file2_wb.save(file1_path)
file2_wb.close()
temp_wb.close()

# Next, we open up a series of files that contain data that link back to the file updated in the previous step. Because
# these files get uploaded to the cloud each day, we have to open them for them to refresh, and save the updated values
file3_path = r'file3_path'  # Define the path
file3_wb = load_workbook(file3_path, keep_links=True)  # Create an instance of our file while preserving links to data
file3_wb.save(file3_path)  # Save the file with new data
file3_wb.close()  # Close the file

# The next phase of the reporting process requires us to generate pdf's of the data that was sent earlier. A previously
# constructed MACRO exists that takes some files saved off earlier and generates pdf's for each and saves them. I repeat
# the code below several times, however, this is meant to show how a macro is accessed via Python.
file4_path = r'file4_path'  # Define the file path
xl = Dispatch('Excel.Application')  # Create an instance of the Excel application
xl.Visible = True  # Visibly see the application
xl.AskToUpdateLinks = False  # Removes an annoying popup

# The following block creates an instance of the file, runs the PDF generating MACRO, and closes the file without saving
file4_wb = xl.Workbooks.Open(file4_path)
xl.Application.Run('Module#.Macro_Name')
file4_wb.Close(SaveChanges=False)

# Throughout the code, I place these to help with stability as a code that runs too fast can be problematic
time.sleep(2)

# The last files that need to get updated are linked to data from previous days. As a result, I created a "mirror" file
# that updates the links by placing the current business day, prior business day, and the 2-day prior business day. The
# "mirror" contains a MACRO that I wrote where the file link dates are changed to reflect the updated day-over-day time
# period. Once the "mirror" file is updated, I simply use the copy/paste method from earlier to update the file that is
# uploaded to the cloud
file5_path = r'file5_path'  # Define the file path
file6_path = r'file6_path'  # Define the file path

# Create an instance of the file
file5_wb = xl.Workbooks.Open(file5_path)
file5_ws = xl.Worksheets('Worksheet_Name')

# Place the dates input by the user from the beginning of the program into cells of the file.
file5_ws.Cells(123, 456).Value = correct_date
file5_ws.Cells(124, 456).Value = old_date
file5_ws.Cells(125, 456).Value = older_date

# The MACRO is then run to change the links of the files to point them to the correct data dates for day-over-day
# measurements. The file is then saved and closed.
xl.Application.Run('Module#.Macro_Name')
file5_wb.Close(SaveChanges=True)

# Define the path of the "main" file that get's uploaded to the cloud
file6_wb = load_workbook(file6_path)
file5_wb = load_workbook(file5_path)

# Similar to the process above, we use the following method to update the values in the "main" file with the values
# from the "mirror" file updated in the previous step
file6_ws = file6_wb.active
file5_ws = file5_wb.active
colLetter = [temp_ws.cell(row=i, column=123).value for i in range(1, 123456)]  # List comprehension technique from above
for i, value in enumerate(colLetter):
    file2_ws.cell(row=123 + i, column=123, value=value)

# Finally, we run a previously created BAT file that automatically uploads all updated files to the cloud
batch_file_path = r'bat_file_path'
p = Popen(batch_file_path, shell=True)
stdout, stderr = p.communicate()

# In conclusion, from start to finish, this program obtains new data, updates all files containing outdated data, and
# uploads all the new data to the cloud. Again, while I have tried to be as specific as possible, I also want to ensure
# that the company is protected from having any of their process shared. I would be happy to go into more detail on what
# each step does with any recruiter or hiring manager that is interested.
